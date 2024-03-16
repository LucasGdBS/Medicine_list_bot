from openpyxl import Workbook

def generate_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Remedios"

    ws['A1'] = 'Nome'
    ws['B1'] = 'Miligramas'
    ws['C1'] = 'Comprimidos'
    ws['D1'] = 'Drogasil'
    ws['E1'] = 'Pague Menos'

    wb.save('remedios.xlsx')
    print('Arquivo gerado com sucesso')

generate_xlsx()
